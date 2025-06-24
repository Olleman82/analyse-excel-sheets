# analyze_excel.py
#
# Detta skript analyserar alla Excel-filer (.xlsx) i den aktuella mappen.
# Det skapar en rapport i Markdown-format (`Excel_Analysis_Report.md`) som
# dokumenterar varje kolumns namn, format och genererar representativ,
# anonymiserad fejkdata.
#
# Krav: pip install pandas openpyxl Faker

import os
import re
from collections import Counter
from datetime import datetime

import numpy as np
import pandas as pd
from faker import Faker
from openpyxl import load_workbook

# Initiera Faker för svensk data
fake = Faker("sv_SE")

# --- Konfiguration ---
SAMPLE_SIZE = 50  # Antal rader att analysera för att gissa datatyp
CATEGORICAL_THRESHOLD = 15  # Max antal unika värden för att ses som "kategorisk"
OUTPUT_FILENAME = "Excel_Analysis_Report.md"
EXAMPLE_COUNT = 5 # Antal exempel att visa i rapporten

# --- Hjälpfunktioner ---

def get_format_info(cell):
    """Hämtar formateringsinformation från en openpyxl-cell."""
    if not cell:
        return "N/A"
    
    info = []
    if cell.font and cell.font.bold:
        info.append("Fet")
    if cell.alignment and cell.alignment.wrapText:
        info.append("Radbruten")
    
    num_format = cell.number_format
    if num_format and num_format not in ["General", None]:
        info.append(f'Format: "{num_format}"')

    return ", ".join(info) if info else "Standard"


def generate_fake_data(series, col_name):
    """Genererar fejkdata baserat på en pandas-series och kolumnnamn."""
    series = series.dropna().astype(str)
    col_name_lower = col_name.lower()

    # Prioriterad ordning för att gissa datatyp från kolumnnamn
    if "namn" in col_name_lower or "kund" in col_name_lower:
        return fake.name()
    if "företag" in col_name_lower:
        return fake.company()
    if "epost" in col_name_lower or "email" in col_name_lower:
        return fake.email()
    if "telefon" in col_name_lower or "mobil" in col_name_lower:
        return fake.phone_number()
    if "adress" in col_name_lower:
        return fake.address().replace("\n", ", ")
    if "personnummer" in col_name_lower:
        return fake.ssn()
    if "nummer" in col_name_lower or "id" in col_name_lower:
        return str(fake.random_number(digits=6, fix_len=True))
    if "datum" in col_name_lower:
        return fake.date()

    # Gissa datatyp från innehåll
    if pd.api.types.is_numeric_dtype(series.infer_objects().dtype):
        return str(round(np.random.uniform(series.min(), series.max()), 2))
    if pd.api.types.is_datetime64_any_dtype(series.infer_objects().dtype):
        return fake.date()

    # Fallback till generisk text
    return fake.word()


def analyze_column(series, col_name):
    """Analyserar en kolumn och returnerar en sammanfattning."""
    
    # 1. Analysera datatyper och proportioner
    total_rows = len(series)
    if total_rows == 0:
        return "Tom kolumn", [], "Ingen data"
        
    type_counts = Counter()
    for val in series:
        if pd.isna(val) or str(val).strip() == "":
            type_counts["Tom"] += 1
        elif isinstance(val, (int, float, np.number)):
            type_counts["Tal"] += 1
        elif isinstance(val, (datetime, pd.Timestamp)):
            type_counts["Datum"] += 1
        else:
            type_counts["Text"] += 1

    type_summary = ", ".join(
        f"{t} ({c/total_rows:.0%})" for t, c in type_counts.items()
    )

    # 2. Analysera unikhet för att bestämma fejkningsmetod
    non_empty_series = series.dropna()
    unique_values = non_empty_series.unique()
    
    # 3. Generera exempeldata
    examples = []
    is_categorical = (len(unique_values) <= CATEGORICAL_THRESHOLD and 
                      len(unique_values) > 0 and 
                      pd.api.types.is_string_dtype(non_empty_series.dtype))

    if is_categorical:
        # Kategorisk data: Visa de faktiska unika värdena (inte känsligt)
        analysis_note = f"Kategorisk ({len(unique_values)} unika värden)"
        examples = list(unique_values[:EXAMPLE_COUNT])
    else:
        # Känslig eller varierande data: Generera fejkdata
        analysis_note = "Varierande/Känslig data"
        for _ in range(EXAMPLE_COUNT):
            examples.append(generate_fake_data(non_empty_series, col_name))
            
    # Återspegla proportionen av tomma värden
    num_empty_to_add = int(type_counts.get("Tom", 0) / total_rows * EXAMPLE_COUNT)
    examples = examples[num_empty_to_add:] # Ta bort några exempel
    for _ in range(num_empty_to_add):
        examples.append("(tom)") # Lägg till "(tom)" för att visa att det finns

    return type_summary, examples, analysis_note


# --- Huvudlogik ---

def main():
    """Huvudfunktion för att köra analysen och skapa rapporten."""
    
    # Hitta alla Excel-filer
    excel_files = [f for f in os.listdir('.') if f.endswith('.xlsx') and not f.startswith('~')]
    
    if not excel_files:
        print("Hittade inga .xlsx-filer i denna mapp.")
        return
        
    print(f"Hittade {len(excel_files)} Excel-fil(er). Analyserar...")
    
    # Skapa rapportfilen
    with open(OUTPUT_FILENAME, "w", encoding="utf-8") as report:
        report.write(f"# Excel-filsanalys\n\n")
        report.write(f"Rapport genererad: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n\n")
        
        for filename in excel_files:
            print(f"  - Analyserar fil: {filename}")
            report.write(f"## Fil: `{filename}`\n\n")
            
            try:
                xls = pd.ExcelFile(filename)
                wb = load_workbook(filename, read_only=True, data_only=True)
            except Exception as e:
                report.write(f"**Kunde inte öppna filen.** Fel: {e}\n\n")
                continue

            for sheet_name in xls.sheet_names:
                report.write(f"### Blad: `{sheet_name}`\n\n")
                
                try:
                    df = pd.read_excel(xls, sheet_name=sheet_name)
                    ws = wb[sheet_name]
                except Exception as e:
                    report.write(f"**Kunde inte läsa bladet.** Fel: {e}\n\n")
                    continue
                    
                if df.empty:
                    report.write("*(Detta blad är tomt.)*\n\n")
                    continue
                
                # Skriv tabellhuvud
                report.write("| Kolumnnamn | Datatyper (andel) | Format (Rubrik) | Format (Data) | Analys/Typ | Fejkade Exempeldata |\n")
                report.write("|------------|-------------------|-----------------|---------------|------------|---------------------|\n")

                for i, col in enumerate(df.columns):
                    type_summary, examples, analysis_note = analyze_column(df[col].head(SAMPLE_SIZE), col)
                    
                    header_cell = ws.cell(row=1, column=i+1)
                    first_data_cell = ws.cell(row=2, column=i+1)
                    
                    header_format = get_format_info(header_cell)
                    data_format = get_format_info(first_data_cell)

                    # Skriv rad i rapporten
                    report.write(
                        f"| `{col}` "
                        f"| {type_summary} "
                        f"| {header_format} "
                        f"| {data_format} "
                        f"| {analysis_note} "
                        f"| `{repr(examples)}` |\n"
                    )
                report.write("\n")
